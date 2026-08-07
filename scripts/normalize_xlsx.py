"""
School Bus Routes — normalise the school intake spreadsheet (xlsx) into the
pipeline's CSV format.

The school's spreadsheet columns (學號/姓名/School 學校/班別/住址 Address/聯絡電話)
map onto the pipeline's format (student_id/name/name_en/school/class_year/
address/dropoff_address/district/contact_phone/contact_name). Fields the
spreadsheet does not carry are left blank rather than invented:
  - name_en            (English name)     -> "" (sheets are Chinese-only)
  - dropoff_address    (PM drop-off)      -> "" (reuses the home coordinates)
  - district           (Chinese district) -> best-effort from known area names,
                                             else "" (AMap geocodes with HK)
  - contact_name       (guardian)         -> ""

Usage:
  python normalize_xlsx.py <in.xlsx> [school] [--sample N] [--out out.csv]

  With [school]: only that school's students are kept.
  With --sample N: pick N students spread across areas (round-robin by area),
  preferring a balanced geographic mix for a demo route. Deterministic
  (lowest 學號 first within each area).
"""
import argparse
import collections
import csv
import random
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

# English area tokens found in the address -> Chinese district (for the AMap
# `city` hint and the mock-geocode fallback). Unknown areas are left blank.
AREA_DISTRICT = {
    "Sha Tin": "沙田", "Tai Wai": "沙田", "Ma On Shan": "沙田",
    "Fo Tan": "沙田", "Siu Lek Yuen": "沙田", "Shek Mun": "沙田",
    "Yuen Wo": "沙田", "City One": "沙田", "Shui Chuen O": "沙田",
    "Tsuen Wan": "荃灣", "Kowloon": "九龍", "Tsing Yi": "葵青",
}

# English district suffix -> Chinese, used when rewriting addresses to Chinese
DISTRICT_EN2CN = {
    "Sha Tin": "沙田", "Tai Wai": "大圍", "Ma On Shan": "馬鞍山",
    "Fo Tan": "火炭", "Siu Lek Yuen": "小瀝源", "Shek Mun": "石門",
    "Yuen Wo": "源禾", "Tsuen Wan": "荃灣", "Tsing Yi": "青衣",
    "Kowloon": "九龍", "Hung Hom": "紅磡", "Kowloon Tong": "九龍塘",
    "Kowloon City": "九龍城", "Ho Man Tin": "何文田", "Mong Kok": "旺角",
}


def to_chinese_address(raw):
    """
    Rewrite a bilingual address like
    "Bayshore Towers 海濤居 18/F #04, Ma On Shan"
    into a Chinese-only form for better AMap geocoding:
    "馬鞍山海濤居18樓4室".
    Returns (chinese_address, chinese_district).
    """
    import re
    addr = (raw or "").strip()
    district = ""
    low = addr.lower()
    for d_en, d_cn in DISTRICT_EN2CN.items():
        if low.endswith(d_en.lower()):
            district = d_cn
            addr = addr[: len(addr) - len(d_en)].rstrip(" ,")
            break
    # floor/unit: 18/F #04 -> 18樓4室 ; 18/F -> 18樓 ; #04 -> 4室
    addr = re.sub(r"(\d+)\s*/\s*F\s*#\s*(\d+)", r"\1樓\2室", addr)
    addr = re.sub(r"(\d+)\s*/\s*F", r"\1樓", addr)
    addr = re.sub(r"#\s*(\d+)", r"\1室", addr)
    # strip the leading English estate name, keep the Chinese + floor
    addr = re.sub(r"^[A-Za-z0-9 &'\-\.]+", "", addr).strip()
    addr = re.sub(r"\s+", "", addr)
    full = (district + addr).strip()
    return (full or raw or "").strip(), district


def read_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl not installed — run: ../.venv/Scripts/pip install openpyxl")
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = [str(c).strip() if c else "" for c in rows[0]]
    return header, [r for r in rows[1:] if r and r[0]]


def area_of(address):
    """Chinese district for this address, from a trailing English area name."""
    a = address or ""
    for token, district in AREA_DISTRICT.items():
        if token.lower() in a.lower():
            return district
    return ""


def to_pipeline_row(raw, header):
    idx = {h: i for i, h in enumerate(header)}
    def get(*names):
        for n in names:
            if n in idx and idx[n] is not None:
                v = raw[idx[n]]
                if v is not None and str(v).strip():
                    return str(v).strip()
        return ""
    address = get("住址 Address", "address", "住址")
    cn_address, district = to_chinese_address(address)
    return {
        "student_id":      get("學號", "student_id", "學號/Student ID"),
        "name":            get("姓名", "name"),
        "name_en":         get("English Name", "name_en"),
        "school":          get("School 學校", "school", "School"),
        "class_year":      get("班別", "class_year", "Class"),
        "address":         cn_address,
        "dropoff_address": get("下車地址", "dropoff_address"),
        "district":        district,
        "contact_phone":   get("聯絡電話", "contact_phone", "電話"),
        "contact_name":    get("監護人", "contact_name", "聯絡人"),
    }


def pick_sample(rows, n):
    """Round-robin across areas (deterministic, lowest 學號 first) for a
    geographically balanced demo set."""
    by_area = collections.defaultdict(list)
    for r in rows:
        by_area[r["district"] or "其他"].append(r)
    for k in by_area:
        by_area[k].sort(key=lambda r: r["student_id"])
    order = sorted(by_area, key=lambda k: -len(by_area[k]))
    picked, iters = [], 0
    while len(picked) < n and iters < n * 2:
        iters += 1
        for k in order:
            if by_area[k]:
                picked.append(by_area[k].pop(0))
            if len(picked) >= n:
                break
    return picked[:n]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("in_xlsx")
    ap.add_argument("school", nargs="?", default="")
    ap.add_argument("--sample", type=int, default=0)
    ap.add_argument("--out", default="")
    args = ap.parse_args()

    header, rows = read_xlsx(args.in_xlsx)
    students = [to_pipeline_row(r, header) for r in rows if to_pipeline_row(r, header)["student_id"]]

    schools = collections.Counter(s["school"] for s in students)
    print(f"Loaded {len(students)} students: {dict(schools)}")

    if args.school:
        students = [s for s in students if s["school"] == args.school]
        print(f"Filtered to {len(students)} students at {args.school}")

    if args.sample:
        students = pick_sample(students, args.sample)
        print(f"Sample: {len(students)} students "
              f"(districts: {dict(collections.Counter(s['district'] or '其他' for s in students))})")

    out = Path(args.out) if args.out else (
        Path(args.in_xlsx).with_suffix(".csv").name.replace("_sample", "")
    )
    out_path = Path(__file__).resolve().parent.parent / "data" / out
    fieldnames = ["student_id", "name", "name_en", "school", "class_year",
                  "address", "dropoff_address", "district", "contact_phone", "contact_name"]
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(students)
    print(f"Wrote {len(students)} rows -> {out_path}")


if __name__ == "__main__":
    main()
